"""Reporte financiero exportable a PDF (ReportLab) y Excel (openpyxl) — RF-17."""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen.canvas import Canvas

PRIMARIO = colors.HexColor("#00526d")
GRIS = colors.HexColor("#40484d")
GRIS_CLARO = colors.HexColor("#70787e")
LINEA = colors.HexColor("#bfc8ce")
VERDE = colors.HexColor("#00573d")
ROJO = colors.HexColor("#ba1a1a")
FONDO = colors.HexColor("#e4f7f9")


def _fmt_ars(valor) -> str:
    entero, dec = f"{Decimal(str(valor)):,.2f}".split(".")
    return f"$ {entero.replace(',', '.')},{dec}"


def _fecha(iso: str) -> str:
    return datetime.fromisoformat(iso).strftime("%d/%m/%Y")


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

class _PaginadorPdf:
    """Canvas con encabezado, pie y salto de página automático."""

    def __init__(self, buffer: BytesIO, titulo: str, subtitulo: str):
        self.pdf = Canvas(buffer, pagesize=A4)
        self.ancho, self.alto = A4
        self.margen = 18 * mm
        self.titulo = titulo
        self.subtitulo = subtitulo
        self.y = 0.0
        self._encabezado()

    def _encabezado(self) -> None:
        from app.services.branding import GESCOM_LOGO, dibujar_logo

        pdf, margen = self.pdf, self.margen
        y = self.alto - self.margen
        dibujar_logo(pdf, GESCOM_LOGO, margen, y + 8, 11 * mm)
        pdf.setFillColor(GRIS)
        pdf.setFont("Helvetica-Bold", 13)
        pdf.drawRightString(self.ancho - margen, y, self.titulo)
        pdf.setFont("Helvetica", 9)
        pdf.setFillColor(GRIS_CLARO)
        pdf.drawRightString(self.ancho - margen, y - 13, self.subtitulo)
        y -= 26
        pdf.setStrokeColor(LINEA)
        pdf.line(margen, y, self.ancho - margen, y)
        self.y = y - 20

    def _pie(self) -> None:
        self.pdf.setFont("Helvetica", 8)
        self.pdf.setFillColor(GRIS_CLARO)
        self.pdf.drawCentredString(
            self.ancho / 2, 12 * mm,
            "Documento generado por GESCOM — Sistema de Gestión de Reservas",
        )

    def asegurar(self, alto_necesario: float) -> None:
        if self.y - alto_necesario < 22 * mm:
            self._pie()
            self.pdf.showPage()
            self._encabezado()

    def seccion(self, texto: str) -> None:
        self.asegurar(30)
        self.pdf.setFont("Helvetica-Bold", 12)
        self.pdf.setFillColor(PRIMARIO)
        self.pdf.drawString(self.margen, self.y, texto)
        self.y -= 18

    def cerrar(self) -> None:
        self._pie()
        self.pdf.showPage()
        self.pdf.save()


def generar_pdf_financiero(data: dict) -> bytes:
    """`data` es el dict de finanzas.reporte (resumen + desglose + ocupación + transacciones)."""
    buffer = BytesIO()
    periodo = f"Período: {_fecha(data['desde'])} al {_fecha(data['hasta'])}"
    doc = _PaginadorPdf(buffer, "Reporte Financiero", periodo)
    pdf, margen, ancho = doc.pdf, doc.margen, doc.ancho
    kpis = data["kpis"]

    # --- KPIs ---
    doc.seccion("Indicadores del período")
    alto_caja = 60
    doc.asegurar(alto_caja + 10)
    pdf.setFillColor(FONDO)
    pdf.roundRect(margen, doc.y - alto_caja + 14, ancho - 2 * margen, alto_caja, 6, stroke=0, fill=1)

    def kpi(etiqueta: str, valor: str, x: float, color=GRIS) -> None:
        pdf.setFont("Helvetica-Bold", 7.5)
        pdf.setFillColor(GRIS_CLARO)
        pdf.drawString(x, doc.y, etiqueta.upper())
        pdf.setFont("Helvetica-Bold", 12)
        pdf.setFillColor(color)
        pdf.drawString(x, doc.y - 15, valor)

    paso = (ancho - 2 * margen - 12 * mm) / 4
    x0 = margen + 6 * mm
    kpi("Ingresos", _fmt_ars(kpis["ingresos_ars"]), x0, VERDE)
    kpi("Egresos", _fmt_ars(kpis["egresos_ars"]), x0 + paso, ROJO)
    balance = kpis["balance_ars"]
    kpi("Balance", _fmt_ars(balance), x0 + 2 * paso, VERDE if balance >= 0 else ROJO)
    kpi("Ocupación promedio", f"{kpis['ocupacion']} %", x0 + 3 * paso, PRIMARIO)
    doc.y -= 34
    pdf.setFont("Helvetica", 8.5)
    pdf.setFillColor(GRIS_CLARO)
    pdf.drawString(
        x0,
        doc.y,
        f"Pendiente de cobro: {_fmt_ars(kpis['pendiente_cobro_ars'])}   ·   "
        f"Ingresos acumulados del año: {_fmt_ars(kpis['ingresos_anio_ars'])}   ·   "
        f"Tipo de cambio: {_fmt_ars(data['tipo_cambio'])} /USD",
    )
    doc.y -= 26

    # --- Desglose por categoría ---
    doc.seccion("Resumen por categoría")

    def linea_categoria(nombre: str, monto, signo: str, color) -> None:
        doc.asegurar(16)
        pdf.setFont("Helvetica", 9.5)
        pdf.setFillColor(GRIS)
        pdf.drawString(margen + 4 * mm, doc.y, nombre)
        pdf.setFillColor(color)
        pdf.drawRightString(ancho - margen, doc.y, f"{signo} {_fmt_ars(monto)}")
        doc.y -= 14

    desglose = data["desglose"]
    linea_categoria("Cobros de reservas (alojamiento)", desglose["ingresos"]["cobros_reservas_ars"], "+", VERDE)
    for item in desglose["ingresos"]["movimientos"]:
        linea_categoria(f"Otros ingresos — {item['categoria']}", item["total_ars"], "+", VERDE)
    for item in desglose["egresos"]["movimientos"]:
        linea_categoria(f"Egresos — {item['categoria']}", item["total_ars"], "-", ROJO)
    doc.asegurar(20)
    pdf.setStrokeColor(LINEA)
    pdf.line(margen, doc.y + 4, ancho - margen, doc.y + 4)
    doc.y -= 8
    pdf.setFont("Helvetica-Bold", 10.5)
    pdf.setFillColor(VERDE if balance >= 0 else ROJO)
    pdf.drawString(margen + 4 * mm, doc.y, "Balance del período")
    pdf.drawRightString(ancho - margen, doc.y, _fmt_ars(balance))
    doc.y -= 28

    # --- Ocupación por unidad ---
    doc.seccion("Ocupación por unidad")
    ocupacion = data["ocupacion"]
    encabezados = ["UNIDAD", "NOCHES OCUP.", "NOCHES PERÍODO", "OCUPACIÓN", "RESERVAS", "INGRESOS"]
    posiciones = [margen, margen + 62 * mm, margen + 92 * mm, margen + 122 * mm, margen + 145 * mm]
    doc.asegurar(30)
    pdf.setFont("Helvetica-Bold", 7.5)
    pdf.setFillColor(GRIS_CLARO)
    for texto, x in zip(encabezados[:-1], posiciones):
        pdf.drawString(x, doc.y, texto)
    pdf.drawRightString(ancho - margen, doc.y, encabezados[-1])
    doc.y -= 4
    pdf.setStrokeColor(LINEA)
    pdf.line(margen, doc.y, ancho - margen, doc.y)
    doc.y -= 13
    for unidad in ocupacion["unidades"]:
        doc.asegurar(16)
        pdf.setFont("Helvetica", 9)
        pdf.setFillColor(GRIS)
        nombre = unidad["nombre"] + ("" if unidad["activo"] else " (inactivo)")
        pdf.drawString(posiciones[0], doc.y, nombre[:38])
        pdf.drawString(posiciones[1], doc.y, str(unidad["noches_ocupadas"]))
        pdf.drawString(posiciones[2], doc.y, str(unidad["noches_disponibles"]))
        pdf.drawString(posiciones[3], doc.y, f"{unidad['porcentaje']} %")
        pdf.drawString(posiciones[4], doc.y, str(unidad["reservas"]))
        pdf.drawRightString(ancho - margen, doc.y, _fmt_ars(unidad["ingresos_ars"]))
        doc.y -= 14
    doc.asegurar(18)
    pdf.setFont("Helvetica-Bold", 9)
    pdf.setFillColor(PRIMARIO)
    pdf.drawString(
        margen, doc.y,
        f"Promedio general: {ocupacion['ocupacion_promedio']} % "
        f"({ocupacion['noches_ocupadas']} noches sobre {ocupacion['unidades_activas']} unidades activas)",
    )
    doc.y -= 26

    # --- Transacciones ---
    doc.seccion("Detalle de transacciones")
    pos_t = [margen, margen + 22 * mm, margen + 95 * mm, margen + 112 * mm, margen + 140 * mm]
    doc.asegurar(30)
    pdf.setFont("Helvetica-Bold", 7.5)
    pdf.setFillColor(GRIS_CLARO)
    for texto, x in zip(["FECHA", "DESCRIPCIÓN", "TIPO", "CATEGORÍA", "DEPTO"], pos_t):
        pdf.drawString(x, doc.y, texto)
    pdf.drawRightString(ancho - margen, doc.y, "MONTO (ARS)")
    doc.y -= 4
    pdf.setStrokeColor(LINEA)
    pdf.line(margen, doc.y, ancho - margen, doc.y)
    doc.y -= 13
    if not data["transacciones"]:
        pdf.setFont("Helvetica-Oblique", 9)
        pdf.setFillColor(GRIS_CLARO)
        pdf.drawString(margen, doc.y, "Sin transacciones en el período.")
        doc.y -= 14
    for fila in data["transacciones"]:
        doc.asegurar(16)
        es_ingreso = fila["tipo"] == "INGRESO"
        pdf.setFont("Helvetica", 8.5)
        pdf.setFillColor(GRIS)
        pdf.drawString(pos_t[0], doc.y, _fecha(fila["fecha"]))
        pdf.drawString(pos_t[1], doc.y, fila["descripcion"][:48])
        pdf.setFillColor(VERDE if es_ingreso else ROJO)
        pdf.drawString(pos_t[2], doc.y, "Ingreso" if es_ingreso else "Egreso")
        pdf.setFillColor(GRIS)
        pdf.drawString(pos_t[3], doc.y, (fila["categoria"] or "-")[:18])
        pdf.drawString(pos_t[4], doc.y, (fila["departamento"] or "Global")[:16])
        signo = "" if es_ingreso else "- "
        pdf.drawRightString(ancho - margen, doc.y, f"{signo}{_fmt_ars(fila['monto_ars'])}")
        doc.y -= 13

    doc.cerrar()
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="00526D")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITULO_FONT = Font(bold=True, size=14, color="00526D")
_BORDE = Border(bottom=Side(style="thin", color="BFC8CE"))
_ARS = '"$" #,##0.00'


def _encabezar(ws, fila: int, encabezados: list[str]) -> None:
    for col, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila, column=col, value=texto)
        celda.fill = _HEADER_FILL
        celda.font = _HEADER_FONT
        celda.alignment = Alignment(horizontal="center")


def _autoancho(ws, anchos: dict[int, int]) -> None:
    for col, ancho in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = ancho


def generar_excel_financiero(data: dict) -> bytes:
    kpis = data["kpis"]
    desglose = data["desglose"]
    wb = Workbook()

    # --- Hoja Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "GESCOM — Reporte Financiero"
    ws["A1"].font = _TITULO_FONT
    ws["A2"] = f"Período: {_fecha(data['desde'])} al {_fecha(data['hasta'])}"
    ws["A3"] = f"Tipo de cambio de referencia: $ {data['tipo_cambio']:,.2f} /USD"

    filas_resumen = [
        ("Ingresos del período (ARS)", kpis["ingresos_ars"]),
        ("Egresos del período (ARS)", kpis["egresos_ars"]),
        ("Balance del período (ARS)", kpis["balance_ars"]),
        ("Ingresos acumulados del año (ARS)", kpis["ingresos_anio_ars"]),
        ("Pendiente de cobro (ARS)", kpis["pendiente_cobro_ars"]),
        ("Pendiente de cobro (USD)", kpis["pendiente_cobro_usd"]),
        ("Ocupación promedio (%)", kpis["ocupacion"]),
        ("Unidades activas", kpis["unidades_activas"]),
    ]
    _encabezar(ws, 5, ["Indicador", "Valor"])
    for i, (nombre, valor) in enumerate(filas_resumen, start=6):
        ws.cell(row=i, column=1, value=nombre).border = _BORDE
        celda = ws.cell(row=i, column=2, value=valor)
        celda.border = _BORDE
        if "ARS" in nombre or "USD" in nombre:
            celda.number_format = _ARS

    fila = len(filas_resumen) + 7
    ws.cell(row=fila, column=1, value="Resumen por categoría").font = Font(bold=True, size=12)
    fila += 1
    _encabezar(ws, fila, ["Concepto", "Tipo", "Total (ARS)"])
    fila += 1

    def linea(concepto: str, tipo: str, total) -> None:
        nonlocal fila
        ws.cell(row=fila, column=1, value=concepto).border = _BORDE
        ws.cell(row=fila, column=2, value=tipo).border = _BORDE
        celda = ws.cell(row=fila, column=3, value=total)
        celda.number_format = _ARS
        celda.border = _BORDE
        fila += 1

    linea("Cobros de reservas (alojamiento)", "Ingreso", desglose["ingresos"]["cobros_reservas_ars"])
    for item in desglose["ingresos"]["movimientos"]:
        linea(item["categoria"], "Ingreso", item["total_ars"])
    for item in desglose["egresos"]["movimientos"]:
        linea(item["categoria"], "Egreso", item["total_ars"])
    celda = ws.cell(row=fila, column=1, value="Balance del período")
    celda.font = Font(bold=True)
    celda_total = ws.cell(row=fila, column=3, value=kpis["balance_ars"])
    celda_total.number_format = _ARS
    celda_total.font = Font(bold=True)
    _autoancho(ws, {1: 38, 2: 12, 3: 18})

    # --- Hoja Transacciones ---
    ws = wb.create_sheet("Transacciones")
    _encabezar(
        ws, 1,
        ["Fecha", "Descripción", "Tipo", "Categoría", "Departamento", "Monto", "Moneda", "Monto (ARS)"],
    )
    for i, t in enumerate(data["transacciones"], start=2):
        ws.cell(row=i, column=1, value=datetime.fromisoformat(t["fecha"]).date())
        ws.cell(row=i, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=i, column=2, value=t["descripcion"])
        ws.cell(row=i, column=3, value="Ingreso" if t["tipo"] == "INGRESO" else "Egreso")
        ws.cell(row=i, column=4, value=t["categoria"] or "-")
        ws.cell(row=i, column=5, value=t["departamento"] or "Global")
        ws.cell(row=i, column=6, value=t["monto"]).number_format = "#,##0.00"
        ws.cell(row=i, column=7, value=t["moneda"])
        monto_ars = t["monto_ars"] if t["tipo"] == "INGRESO" else -t["monto_ars"]
        ws.cell(row=i, column=8, value=monto_ars).number_format = _ARS
    ws.auto_filter.ref = f"A1:H{max(1, len(data['transacciones']) + 1)}"
    _autoancho(ws, {1: 12, 2: 48, 3: 10, 4: 18, 5: 20, 6: 14, 7: 9, 8: 16})

    # --- Hoja Ocupación ---
    ws = wb.create_sheet("Ocupación")
    ocupacion = data["ocupacion"]
    _encabezar(
        ws, 1,
        ["Unidad", "Noches ocupadas", "Noches del período", "Ocupación (%)", "Reservas", "Ingresos (ARS)"],
    )
    for i, u in enumerate(ocupacion["unidades"], start=2):
        ws.cell(row=i, column=1, value=u["nombre"] + ("" if u["activo"] else " (inactivo)"))
        ws.cell(row=i, column=2, value=u["noches_ocupadas"])
        ws.cell(row=i, column=3, value=u["noches_disponibles"])
        ws.cell(row=i, column=4, value=u["porcentaje"])
        ws.cell(row=i, column=5, value=u["reservas"])
        ws.cell(row=i, column=6, value=u["ingresos_ars"]).number_format = _ARS
    fila = len(ocupacion["unidades"]) + 2
    celda = ws.cell(row=fila, column=1, value="Promedio general")
    celda.font = Font(bold=True)
    ws.cell(row=fila, column=4, value=ocupacion["ocupacion_promedio"]).font = Font(bold=True)
    _autoancho(ws, {1: 32, 2: 18, 3: 18, 4: 14, 5: 10, 6: 16})

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
